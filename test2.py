import openpyxl
import json
import io
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell


# excel表格转json文件
def excel_to_json(excel_file, json_file_name):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        value = sheet.cell(1, column + 1).value
        if isinstance(value, str):
            value = value.replace("\n", ' ')
        print(type(value))
        heads.append(value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = parser_merged_cell(sheet, row + 1, column + 1)
            value = cell.value
            if isinstance(value, str):
                value = value.replace("\n", ' ')
            one_line[k] = value
        # print(one_line)
        result.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(result, json_file_name)


# 将json保存为文件
def save_json_file(jd, json_file_name):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    # 把对象转化为json对象
    # indent: 参数根据数据格式缩进显示，读起来更加清晰
    # ensure_ascii = True：默认输出ASCII码，如果把这个该成False, 就可以输出中文。
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()


# 处理合并单元格
def parser_merged_cell(sheet: Worksheet, row, column):
    cell = sheet.cell(row, column)
    # 判断该单元格是否为合并单元格
    if isinstance(cell, MergedCell):
        # 循环查找该单元格所属的合并区域
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                break
    return cell


if '__main__' == __name__:
    excel_to_json(u'error_code2.xlsx', 'result.json')
